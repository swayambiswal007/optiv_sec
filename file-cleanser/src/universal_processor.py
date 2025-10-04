# src/universal_processor.py - Process all file types
import os
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any
import cv2
import numpy as np

from .config import Config

class UniversalFileProcessor:
    """Process any supported file type"""
    
    def __init__(self):
        self.processors = {
            'image': self._process_image,
            'text': self._process_text_file,
            'json': self._process_json_file,
            'xml': self._process_xml_file,
            'excel': self._process_excel_file,
            'pdf': self._process_pdf_file,
            'csv': self._process_csv_file,
        }
    
    def detect_file_type(self, file_path: str) -> str:
        """Detect file type from extension"""
        ext = Path(file_path).suffix.lower()
        
        if ext in Config.SUPPORTED_IMAGE_FORMATS:
            return 'image'
        elif ext in ['.json']:
            return 'json'
        elif ext in ['.xml']:
            return 'xml'
        elif ext in ['.xlsx', '.xls', '.ods']:
            return 'excel'
        elif ext in ['.pdf']:
            return 'pdf'
        elif ext in ['.csv']:
            return 'csv'
        elif ext in Config.SUPPORTED_TEXT_FORMATS:
            return 'text'
        else:
            return 'unknown'
    
    def process_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process any file type"""
        file_type = self.detect_file_type(file_path)
        
        if file_type == 'unknown':
            raise ValueError(f"Unsupported file type: {Path(file_path).suffix}")
        
        processor = self.processors.get(file_type)
        return processor(file_path, sensitive_detector, text_cleaner)
    
    # def _process_image(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
    #     """Process image files with OCR"""
    #     from .text_extractor import TextExtractor
    #     from .image_redactor import ImageRedactor
        
    #     extractor = TextExtractor()
    #     redactor = ImageRedactor()
        
    #     # Extract text and bounding boxes
    #     text, bounding_boxes = extractor.extract_from_image(file_path)
        
    #     result = {
    #         'file_type': 'image',
    #         'text_extracted': text,
    #         'text_length': len(text),
    #         'bounding_boxes': bounding_boxes,
    #         'sensitive_items': [],
    #         'output_files': []
    #     }
        
    #     # Clean text
    #     cleaned_text = text_cleaner.clean_text(text) if text else ""
        
    #     # Detect sensitive data
    #     text_matches = sensitive_detector.detect_in_text(cleaned_text)
    #     result['sensitive_items'] = text_matches
        
    #     # Detect in bounding boxes
    #     sensitive_boxes = []
    #     if bounding_boxes:
    #         sensitive_boxes = sensitive_detector.detect_large_regions_for_redaction(
    #             bounding_boxes, 
    #             self._get_image_shape(file_path)
    #         )
        
    #     # Apply document pattern recognition
    #     doc_type = self._detect_document_type(cleaned_text)
    #     result['document_type'] = doc_type
        
    #     if doc_type and not sensitive_boxes:
    #         # Apply pattern-based redaction for known document types
    #         sensitive_boxes = self._get_document_pattern_regions(file_path, doc_type)
        
    #     # Create outputs
    #     file_stem = Path(file_path).stem
        
    #     # Redacted text file
    #     redacted_text = sensitive_detector.redact_text(cleaned_text, text_matches)
    #     text_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.txt"
    #     with open(text_output, 'w', encoding='utf-8') as f:
    #         f.write(redacted_text)
    #     result['output_files'].append(str(text_output))
        
    #     # Redacted image
    #     if sensitive_boxes or text_matches:
    #         image_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.png"
            
    #         if sensitive_boxes:
    #             redactor.redact_image(file_path, sensitive_boxes, str(image_output))
    #         else:
    #             # Fallback: create redaction zones based on text matches
    #             fallback_boxes = self._create_fallback_boxes(bounding_boxes, text_matches, cleaned_text)
    #             if fallback_boxes:
    #                 redactor.redact_image(file_path, fallback_boxes, str(image_output))
    #             else:
    #                 # Just copy the image
    #                 img = cv2.imread(file_path)
    #                 cv2.imwrite(str(image_output), img)
            
    #         result['output_files'].append(str(image_output))
    #         result['redaction_applied'] = True
    #     else:
    #         # No sensitive data - copy original
    #         img = cv2.imread(file_path)
    #         image_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_processed.png"
    #         cv2.imwrite(str(image_output), img)
    #         result['output_files'].append(str(image_output))
    #         result['redaction_applied'] = False
        
    #     return result
    def _process_image(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        from .text_extractor import TextExtractor
        from .image_redactor import ImageRedactor
    
        extractor = TextExtractor()
        redactor = ImageRedactor()

        # Step 1: OCR
        text, bounding_boxes = extractor.extract_from_image(file_path)

        result = {
            'file_type': 'image',
            'text_extracted': text,
            'text_length': len(text),
            'bounding_boxes': bounding_boxes,
            'sensitive_items': [],
            'output_files': []
        }

        # Step 2: Clean & detect sensitive text
        cleaned_text = text_cleaner.clean_text(text) if text else ""
        text_matches = sensitive_detector.detect_in_text(cleaned_text)
        result['sensitive_items'] = text_matches

        # Step 3: Find bounding boxes for sensitive text only
        sensitive_boxes = []
        for match in text_matches:
            for box in bounding_boxes:
                if match['text'] in box['text']:  # OCR fragment contains sensitive item
                    sensitive_boxes.append(box)

        # Step 4: Save redacted outputs
        file_stem = Path(file_path).stem
        redacted_text = sensitive_detector.redact_text(cleaned_text, text_matches)

        text_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.txt"
        with open(text_output, 'w', encoding='utf-8') as f:
            f.write(redacted_text)
        result['output_files'].append(str(text_output))

        if sensitive_boxes:
            image_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.png"
            redactor.redact_image(file_path, sensitive_boxes, str(image_output))
            result['output_files'].append(str(image_output))
            result['redaction_applied'] = True
        else:
            # No sensitive data detected, just copy original
            img = cv2.imread(file_path)
            image_output = Path(Config.OUTPUT_DIR) / f"{file_stem}_processed.png"
            cv2.imwrite(str(image_output), img)
            result['output_files'].append(str(image_output))
            result['redaction_applied'] = False

        return result

    
    def _process_text_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process plain text files"""
        # Read file
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            text = f.read()
        
        result = {
            'file_type': 'text',
            'text_length': len(text),
            'sensitive_items': [],
            'output_files': []
        }
        
        # Clean and detect
        cleaned_text = text_cleaner.clean_text(text)
        matches = sensitive_detector.detect_in_text(cleaned_text)
        result['sensitive_items'] = matches
        
        # Redact and save
        redacted_text = sensitive_detector.redact_text(cleaned_text, matches)
        
        file_stem = Path(file_path).stem
        output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted{Path(file_path).suffix}"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(redacted_text)
        
        result['output_files'].append(str(output_path))
        result['redaction_applied'] = len(matches) > 0
        
        return result
    
    def _process_json_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process JSON files"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        result = {
            'file_type': 'json',
            'sensitive_items': [],
            'output_files': []
        }
        
        # Recursively redact JSON
        redacted_data, matches = self._redact_json_recursive(data, sensitive_detector)
        result['sensitive_items'] = matches
        
        # Save redacted JSON
        file_stem = Path(file_path).stem
        output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.json"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(redacted_data, f, indent=2 if Config.PRETTY_PRINT_OUTPUT else None, ensure_ascii=False)
        
        result['output_files'].append(str(output_path))
        result['redaction_applied'] = len(matches) > 0
        
        return result
    
    def _process_xml_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process XML files"""
        with open(file_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        result = {
            'file_type': 'xml',
            'sensitive_items': [],
            'output_files': []
        }
        
        # Detect sensitive data in text content
        matches = sensitive_detector.detect_in_text(content)
        result['sensitive_items'] = matches
        
        # Redact
        redacted_content = sensitive_detector.redact_text(content, matches)
        
        # Save
        file_stem = Path(file_path).stem
        output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.xml"
        
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write(redacted_content)
        
        result['output_files'].append(str(output_path))
        result['redaction_applied'] = len(matches) > 0
        
        return result
    
    def _process_excel_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process Excel files"""
        try:
            import openpyxl
            from openpyxl.styles import PatternFill
        except ImportError:
            print("⚠️  openpyxl not installed. Install with: pip install openpyxl")
            return self._process_excel_fallback(file_path, sensitive_detector)
        
        wb = openpyxl.load_workbook(file_path)
        
        result = {
            'file_type': 'excel',
            'sheets_processed': [],
            'sensitive_items': [],
            'output_files': []
        }
        
        # Process each sheet
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            sheet_matches = []
            
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value and isinstance(cell.value, str):
                        # Check cell value for sensitive data
                        cell_matches = sensitive_detector.detect_in_text(str(cell.value))
                        
                        if cell_matches:
                            # Redact the cell
                            cell.value = "[REDACTED]"
                            # Highlight the cell
                            cell.fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                            sheet_matches.extend(cell_matches)
            
            result['sheets_processed'].append({
                'name': sheet_name,
                'sensitive_items_found': len(sheet_matches)
            })
            result['sensitive_items'].extend(sheet_matches)
        
        # Save and close workbook
        wb.save(Path(Config.OUTPUT_DIR) / f"{Path(file_path).stem}_redacted.xlsx")
        result['output_files'].append(str(Path(Config.OUTPUT_DIR) / f"{Path(file_path).stem}_redacted.xlsx"))
        result['redaction_applied'] = len(result['sensitive_items']) > 0
        
        return result
    
    def _process_csv_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process CSV files"""
        import csv
        
        result = {
            'file_type': 'csv',
            'rows_processed': 0,
            'sensitive_items': [],
            'output_files': []
        }
        
        # Read CSV
        rows = []
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            reader = csv.reader(f)
            headers = next(reader, None)
            
            for row in reader:
                redacted_row = []
                for cell in row:
                    matches = sensitive_detector.detect_in_text(str(cell))
                    if matches:
                        redacted_row.append("[REDACTED]")
                        result['sensitive_items'].extend(matches)
                    else:
                        redacted_row.append(cell)
                rows.append(redacted_row)
                result['rows_processed'] += 1
        
        # Save redacted CSV
        file_stem = Path(file_path).stem
        output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.csv"
        
        with open(output_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            if headers:
                writer.writerow(headers)
            writer.writerows(rows)
        
        result['output_files'].append(str(output_path))
        result['redaction_applied'] = len(result['sensitive_items']) > 0
        
        return result
    
    def _process_pdf_file(self, file_path: str, sensitive_detector, text_cleaner) -> Dict:
        """Process PDF files by converting to images"""
        try:
            from pdf2image import convert_from_path
        except ImportError:
            print("⚠️  pdf2image not installed. Install with: pip install pdf2image")
            print("⚠️  Also install poppler: https://github.com/oschwartz10612/poppler-windows/releases/")
            return {'file_type': 'pdf', 'error': 'pdf2image not available'}
        
        result = {
            'file_type': 'pdf',
            'pages_processed': 0,
            'sensitive_items': [],
            'output_files': []
        }
        
        # Convert PDF pages to images
        images = convert_from_path(file_path, dpi=Config.PDF_DPI)
        
        from .text_extractor import TextExtractor
        from .image_redactor import ImageRedactor
        
        extractor = TextExtractor()
        redactor = ImageRedactor()
        
        # Process each page
        file_stem = Path(file_path).stem
        
        for page_num, image in enumerate(images, 1):
            # Save temp image
            temp_path = Path(Config.TEMP_DIR) / f"temp_page_{page_num}.png"
            os.makedirs(Config.TEMP_DIR, exist_ok=True)
            image.save(temp_path)
            
            # Process like an image
            text, bounding_boxes = extractor.extract_from_image(str(temp_path))
            cleaned_text = text_cleaner.clean_text(text)
            matches = sensitive_detector.detect_in_text(cleaned_text)
            
            result['sensitive_items'].extend(matches)
            
            if matches or bounding_boxes:
                sensitive_boxes = sensitive_detector.detect_large_regions_for_redaction(
                    bounding_boxes, 
                    (image.height, image.width, 3)
                )
                
                output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_page_{page_num}_redacted.png"
                
                if sensitive_boxes:
                    redactor.redact_image(str(temp_path), sensitive_boxes, str(output_path))
                else:
                    image.save(output_path)
                
                result['output_files'].append(str(output_path))
            
            # Clean up temp file
            if Config.SECURE_DELETE_TEMP_FILES:
                os.remove(temp_path)
            
            result['pages_processed'] += 1
        
        result['redaction_applied'] = len(result['sensitive_items']) > 0
        
        return result
    
    def _redact_json_recursive(self, data: Any, sensitive_detector, path="") -> Tuple[Any, List]:
        """Recursively redact sensitive data in JSON"""
        matches = []
        
        if isinstance(data, dict):
            redacted = {}
            for key, value in data.items():
                new_path = f"{path}.{key}" if path else key
                redacted_value, value_matches = self._redact_json_recursive(value, sensitive_detector, new_path)
                redacted[key] = redacted_value
                matches.extend(value_matches)
            return redacted, matches
        
        elif isinstance(data, list):
            redacted = []
            for idx, item in enumerate(data):
                new_path = f"{path}[{idx}]"
                redacted_item, item_matches = self._redact_json_recursive(item, sensitive_detector, new_path)
                redacted.append(redacted_item)
                matches.extend(item_matches)
            return redacted, matches
        
        elif isinstance(data, str):
            # Check if string contains sensitive data
            item_matches = sensitive_detector.detect_in_text(data)
            if item_matches:
                matches.extend([{**m, 'json_path': path} for m in item_matches])
                return "[REDACTED]", matches
            return data, matches
        
        else:
            # Numbers, booleans, None
            return data, matches
    
    def _get_image_shape(self, image_path: str) -> Tuple[int, int, int]:
        """Get image dimensions"""
        try:
            img = cv2.imread(image_path)
            return img.shape if img is not None else (0, 0, 0)
        except:
            return (0, 0, 0)
    
    def _detect_document_type(self, text: str) -> str:
        """Detect document type from text content"""
        text_lower = text.lower()
        
        for doc_type, indicators in Config.DOCUMENT_TYPE_INDICATORS.items():
            if any(indicator.lower() in text_lower for indicator in indicators):
                return doc_type
        
        return 'unknown'
    
    # def _get_document_pattern_regions(self, image_path: str, doc_type: str) -> List[Dict]:
    #     """Get redaction regions based on document type"""
    #     img = cv2.imread(image_path)
    #     if img is None:
    #         return []
        
    #     height, width = img.shape[:2]
        
    #     # Define standard regions for known document types
    #     patterns = {
    #         'aadhaar_card': [
    #             {'x': int(width * 0.28), 'y': int(height * 0.38), 'w': int(width * 0.45), 'h': int(height * 0.12), 'type': 'name'},
    #             {'x': int(width * 0.28), 'y': int(height * 0.48), 'w': int(width * 0.35), 'h': int(height * 0.08), 'type': 'dob'},
    #             {'x': int(width * 0.28), 'y': int(height * 0.55), 'w': int(width * 0.25), 'h': int(height * 0.08), 'type': 'gender'},
    #             {'x': int(width * 0.15), 'y': int(height * 0.72), 'w': int(width * 0.55), 'h': int(height * 0.12), 'type': 'aadhaar_number'},
    #             {'x': int(width * 0.72), 'y': int(height * 0.25), 'w': int(width * 0.25), 'h': int(height * 0.45), 'type': 'qr_code'},
    #         ],
    #         'pan_card': [
    #             {'x': int(width * 0.25), 'y': int(height * 0.35), 'w': int(width * 0.50), 'h': int(height * 0.15), 'type': 'name'},
    #             {'x': int(width * 0.25), 'y': int(height * 0.50), 'w': int(width * 0.50), 'h': int(height * 0.10), 'type': 'father_name'},
    #             {'x': int(width * 0.25), 'y': int(height * 0.60), 'w': int(width * 0.35), 'h': int(height * 0.08), 'type': 'dob'},
    #             {'x': int(width * 0.20), 'y': int(height * 0.70), 'w': int(width * 0.40), 'h': int(height * 0.10), 'type': 'pan_number'},
    #         ],
    #         'credit_card': [
    #             {'x': int(width * 0.10), 'y': int(height * 0.40), 'w': int(width * 0.80), 'h': int(height * 0.15), 'type': 'card_number'},
    #             {'x': int(width * 0.10), 'y': int(height * 0.60), 'w': int(width * 0.50), 'h': int(height * 0.10), 'type': 'name'},
    #             {'x': int(width * 0.60), 'y': int(height * 0.60), 'w': int(width * 0.20), 'h': int(height * 0.10), 'type': 'expiry'},
    #             {'x': int(width * 0.80), 'y': int(height * 0.60), 'w': int(width * 0.15), 'h': int(height * 0.10), 'type': 'cvv'},
    #         ],
    #     }
        
    #     return patterns.get(doc_type, [])
    
    def _create_fallback_boxes(self, bounding_boxes: List[Dict], text_matches: List[Dict], full_text: str) -> List[Dict]:
        """Create bounding boxes from text matches when OCR boxes don't contain sensitive data"""
        fallback_boxes = []
        
        for match in text_matches:
            # Find which bounding box contains this text
            for box in bounding_boxes:
                if match['text'] in box['text']:
                    fallback_boxes.append(box)
                    break
        
        return fallback_boxes
        
        # Save redacted workbook
        file_stem = Path(file_path).stem
        output_path = Path(Config.OUTPUT_DIR) / f"{file_stem}_redacted.xlsx"
        wb.save(output_path)
        
        result['output_files'].append(str(output_path))
        result['redaction_applied'] = len(result['sensitive_items']) > 0
        
        return result
    
    def _process_excel_fallback(self, file_path: str, sensitive_detector) -> Dict:
        """Fallback Excel processing using pandas"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("Neither openpyxl nor pandas is installed")
        
        # Read all sheets
        dfs = pd.read_excel(file_path, sheet_name=None)
        
        result = {
            'file_type': 'excel',
            'sheets_processed': [],
            'sensitive_items': [],
            'output_files': []
        }
        
        # Process each sheet
        writer = pd.ExcelWriter(Path(Config.OUTPUT_DIR) / f"{Path(file_path).stem}_redacted.xlsx")
        
        for sheet_name, df in dfs.items():
            sheet_matches = []
            
            # Check each cell
            for col in df.columns:
                for idx in df.index:
                    value = df.at[idx, col]
                    if pd.notna(value) and isinstance(value, str):
                        matches = sensitive_detector.detect_in_text(str(value))
                        if matches:
                            df.at[idx, col] = "[REDACTED]"
                            sheet_matches.extend(matches)
            
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            result['sheets_processed'].append({
                'name': sheet_name,
                'sensitive_items_found': len(sheet_matches)
            })
            result['sensitive_items']